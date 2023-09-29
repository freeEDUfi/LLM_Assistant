import sys
import threading
from kivy.clock import Clock
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.filechooser import FileChooserIconView
from kivy.uix.textinput import TextInput
from kivy.uix.popup import Popup
from docx import Document
from openpyxl import load_workbook
from langchain.llms import LlamaCpp

n_gpu_layers = 35
n_batch = 1550
n_ctx = 2000

llm = LlamaCpp(
    model_path="llama-2-7b-chat.Q4_K_M.gguf",
    n_gpu_layers=n_gpu_layers,
    n_batch=n_batch,
    n_ctx=n_ctx,
    max_tokens=0,
    verbose=True
)

class LLMWidget(BoxLayout):
    def __init__(self, **kwargs):
        super(LLMWidget, self).__init__(**kwargs)
        self.orientation = 'vertical'
        self.text_input = TextInput(size_hint_y=0.8)
        self.add_widget(self.text_input)
        self.button_layout = BoxLayout(size_hint_y=0.2)
        self.load_button = Button(text='Load File')
        self.button_layout.add_widget(self.load_button)
        self.load_button.bind(on_release=self.show_filechooser)
        self.analyze_button = Button(text='Analyze')
        self.button_layout.add_widget(self.analyze_button)
        self.analyze_button.bind(on_release=self.analyze_content)
        self.add_widget(self.button_layout)

    def show_filechooser(self, instance):
        self.filechooser = FileChooserIconView(filters=['*.txt', '*.docx', '*.xlsx'])
        box_layout = BoxLayout(orientation='vertical')
        box_layout.add_widget(self.filechooser)
        select_button = Button(text='Select')
        select_button.bind(on_release=self.load_file)
        box_layout.add_widget(select_button)
        self.popup = Popup(title='Choose a File', content=box_layout, size_hint=(0.9, 0.9))
        self.popup.open()

    def load_file(self, instance):
        try:
            file_path = self.filechooser.selection[0]  # Get the selected file path
            if file_path.endswith('.docx'):  # Check if the file is a Word document
                doc = Document(file_path)  # Load the Word document
                content = "\n".join([p.text for p in doc.paragraphs])  # Extract text from the document
            elif file_path.endswith('.xlsx'):  # Check if the file is an Excel spreadsheet
                workbook = load_workbook(file_path)  # Load the Excel workbook
                sheet = workbook.active  # Get the active sheet
                # Extract text from the sheet, formatted as tab-separated values
                content = "\n".join(["\t".join([str(cell) if cell else "" for cell in row]) for row in
                                     sheet.iter_rows(values_only=True)])
            else:  # Assume the file is a text file
                with open(file_path, 'r') as file:  # Open the file
                    content = file.read()  # Read the content of the file

            # Prefix the content with "Content of the File:" and display it in the text input widget
            self.text_input.text = f"Content of the File:\n{content}"
            self.popup.dismiss()
        except Exception as e:
            self.text_input.text = f"Failed to load file: {e}"

    def analyze_content(self, instance):
        content = self.text_input.text
        self.text_input.text += "\nPlease wait..."
        Clock.schedule_once(lambda dt: self.deferred_analysis(content), 0)

    def deferred_analysis(self, content):
        try:
            full_prompt = "explain this " + content + "explain one by one and make it short avoid cutting off your explanation "
            result = llm(full_prompt)
            result = result.strip()
            result_lines = [line for line in result.split('\n') if line.strip()]
            result = "\n".join(result_lines)
            result = result.replace(full_prompt, "")
            Clock.schedule_once(lambda dt: self.update_ui_with_result(content, result), 0)
        except Exception as e:
            Clock.schedule_once(lambda dt: self.update_ui_with_error(e), 0)

    def update_ui_with_result(self, content, result):
        self.text_input.text = content + f"\n\nModel Result:\n{result}"

    def update_ui_with_error(self, e):
        self.text_input.text += f"\nError during analysis: {e}"
